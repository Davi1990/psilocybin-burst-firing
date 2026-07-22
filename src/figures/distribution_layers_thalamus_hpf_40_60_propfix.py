"""
Distribution Plots with Half-Violin Overlay — 40–60 min Post-Injection
======================================================================
PROP_BURSTING STATISTICAL FIX — see RAW-PROP-1-style finding in
Burst analysis/DEBUG_LOG.md. Only compute_prop_bursting_deltas() differs
from distribution_layers_thalamus_hpf_40_60.py: prop_bursting is now
computed as a SESSION-LEVEL aggregate (fraction of units bursting per
region/session/bin) with an effectively session-only bootstrap (unit_id
set equal to session_id), exactly matching heatmap_hier_bootstrap.py /
heatmap_hpf_thalamic_nuclei.py and the Methods text. The previous version
computed a per-unit fraction-of-time-bursting score and ran it through the
standard two-level (session, unit) hierarchical bootstrap instead.
burst_fraction/burst_rate/mean_burst_size are untouched and identical to
the validated original script.

Hierarchical bootstrap (session → unit) for burst_fraction/burst_rate/
mean_burst_size; session-only (via the unit_id=session_id trick) for
prop_bursting. FDR-corrected per panel. The bootstrap distribution of the
group mean is drawn as a half-violin (drug above the row line, control
below). Point + 95% CI inside each half-violin. Stars + Cohen's d on the
right.

Three plot configurations
-------------------------
  1. Cortical layers (14 rows)
       Prefrontal — L1/2/3, L5, L6
       Somatosensory — L1/2/3, L4, L5, L6
       Visual — L1/2/3, L4, L5, L6
       RSP — L1/2/3, L5, L6
  2. Thalamic functional groups (6 rows)
       Anterior, Higher-order, First-order Somatomotor,
       Intralaminar, First-order Sensory Geniculate, Reticular nucleus
  3. Hippocampal nuclei (5 rows)
       CA1, CA2, CA3, DG, Subicular complex (SUB+ProS)

Metrics
-------
  burst_fraction, burst_rate, mean_burst_size  → absolute delta (effect − baseline) per unit
  prop_bursting                                → absolute change in session-level bursting fraction

Time window: mean across effect_40-45, effect_45-50, effect_50-55, effect_55-60

Usage
-----
  python distribution_layers_thalamus_hpf_40_60_propfix.py \
      --data_dir ../data_extraction --nboots 10000
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import glob
import re
import os
import argparse
import warnings
from collections import Counter, defaultdict

warnings.filterwarnings('ignore')

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable, desc='', **kw):
        if desc:
            print(f"    {desc}...")
        return iterable

try:
    from scipy.stats import gaussian_kde
except ImportError:
    raise ImportError("scipy is required for KDE half-violins")


# ============================================================================
# CONSTANTS
# ============================================================================

CONDITION_COLORS = {
    'Saline':     '#030aa7',
    'Psilocybin': '#a90308',
    'Ketanserin': '#F3CA4B',
}

METRIC_LABELS = {
    'burst_fraction':  'Burst Fraction',
    'burst_rate':      'Burst Rate (s⁻¹)',
    'mean_burst_size': 'Mean Burst Size (spikes)',
    'prop_bursting':   'Prop. Bursting Units',
}

COMPARISONS = [
    ('Psilocybin', 'Saline'),
    ('Ketanserin', 'Saline'),
]

# Effect window — collapse across these 5-min bins
EFFECT_WINDOW = {'effect_40-45', 'effect_45-50', 'effect_50-55', 'effect_55-60'}
WINDOW_LABEL  = '40-60 min post-injection'

# Allen IDs
ISOCORTEX_ID = 315
TH_ID        = 549
OLF_ID       = 698
HPF_ID       = 1089
STR_ID       = 477
RT_ID        = 262
CA_ID        = 375

LAYER_NORMALIZE = {'1': '1/2/3', '2/3': '1/2/3', '6a': '6', '6b': '6'}
LAYER_ORDER     = ['1/2/3', '4', '5', '6']

# ── Plot 1: cortical layer rows ──────────────────────────────────────────────
CORTICAL_AREA_TO_GROUP = {
    'MO': 'Prefrontal',  'ORB': 'Prefrontal', 'ACA': 'Prefrontal',
    'ILA': 'Prefrontal', 'PL':  'Prefrontal', 'FRP': 'Prefrontal',
    'SS':  'Somatosensory',
    'VIS': 'Visual',
    'RSP': 'RSP',
}

CORTICAL_ROWS = [
    ('Prefrontal',    '1/2/3'),
    ('Prefrontal',    '5'),
    ('Prefrontal',    '6'),
    ('Somatosensory', '1/2/3'),
    ('Somatosensory', '4'),
    ('Somatosensory', '5'),
    ('Somatosensory', '6'),
    ('Visual',        '1/2/3'),
    ('Visual',        '4'),
    ('Visual',        '5'),
    ('Visual',        '6'),
    ('RSP',           '1/2/3'),
    ('RSP',           '5'),
    ('RSP',           '6'),
]

def _cortical_key(group, layer):
    return f'{group} — L{layer}'

CORTICAL_ROW_ORDER  = [_cortical_key(g, l) for g, l in CORTICAL_ROWS]
CORTICAL_ROW_LABELS = {k: k for k in CORTICAL_ROW_ORDER}

# ── Plot 2: thalamic functional groups (from heatmap_hpf_thalamic_nuclei.py) ─
THALAMIC_GROUP_MAP = {
    'AD': 'Anterior', 'AV': 'Anterior', 'AMd': 'Anterior', 'AMv': 'Anterior',
    'LD': 'Anterior', 'IAD': 'Anterior',
    'MD': 'Higher-order', 'LP': 'Higher-order', 'PO': 'Higher-order',
    'PoT': 'Higher-order', 'SGN': 'Higher-order',
    'VAL': 'First-order Somatomotor', 'VPL': 'First-order Somatomotor',
    'VPM': 'First-order Somatomotor',
    'CL': 'Intralaminar', 'CM': 'Intralaminar', 'PCN': 'Intralaminar',
    'PF': 'Intralaminar', 'SPFp': 'Intralaminar',
    'LGd': 'First-order Sensory Geniculate', 'LGco': 'First-order Sensory Geniculate',
    'LGip': 'First-order Sensory Geniculate', 'LGsh': 'First-order Sensory Geniculate',
    'LGv': 'First-order Sensory Geniculate', 'IGL': 'First-order Sensory Geniculate',
    'MGd': 'First-order Sensory Geniculate', 'MGm': 'First-order Sensory Geniculate',
    'MGv': 'First-order Sensory Geniculate',
    'RT': 'Reticular nucleus',
}

THALAMIC_GROUP_ORDER = [
    'Anterior', 'Higher-order', 'First-order Somatomotor',
    'Intralaminar', 'First-order Sensory Geniculate', 'Reticular nucleus',
]

# ── Plot 3: hippocampal nuclei ───────────────────────────────────────────────
HPF_NUCLEUS_MERGE = {'ProS': 'Subicular complex', 'SUB': 'Subicular complex'}
HPF_ROW_ORDER = ['CA1', 'CA2', 'CA3', 'DG', 'Subicular complex']


# ============================================================================
# ALLEN SDK SETUP
# ============================================================================

ALLEN_SDK_AVAILABLE = False
structure_tree      = None


class ManualStructureTree:
    def __init__(self, structures):
        self.structures        = structures
        self.structures_by_id  = {s['id']: s for s in structures}
        self.structures_by_acr = defaultdict(list)
        for s in structures:
            acr = s.get('acronym')
            if acr:
                self.structures_by_acr[acr].append(s)

    def get_structures_by_acronym(self, acronyms):
        if isinstance(acronyms, str):
            acronyms = [acronyms]
        out = []
        for a in acronyms:
            out.extend(self.structures_by_acr.get(a, []))
        return out

    def get_structures_by_id(self, ids):
        if isinstance(ids, (int, np.integer)):
            ids = [ids]
        return [self.structures_by_id[i] for i in ids if i in self.structures_by_id]


try:
    print("Loading Allen Brain Atlas structure tree...")
    from allensdk.api.queries.ontologies_api import OntologiesApi
    oapi            = OntologiesApi()
    structure_graph = oapi.get_structures_with_sets([1])
    print(f"  Downloaded {len(structure_graph)} Allen structures")
    try:
        from allensdk.core.structure_tree import StructureTree
        structure_tree = StructureTree(structure_graph)
        if structure_tree.get_structures_by_acronym(['VPM']):
            ALLEN_SDK_AVAILABLE = True
            print("  Allen SDK StructureTree working")
        else:
            raise ValueError("empty result")
    except Exception as e:
        print(f"  StructureTree fallback ({e}); using manual workaround")
        structure_tree = ManualStructureTree(structure_graph)
        if structure_tree.get_structures_by_acronym(['VPM']):
            ALLEN_SDK_AVAILABLE = True
            print("  Manual structure tree working")
        else:
            raise ValueError("manual tree failed too")
except Exception as e:
    print(f"  Allen SDK unavailable: {e}")
    ALLEN_SDK_AVAILABLE = False


def _get_path(structure):
    p = structure.get('structure_id_path', [])
    if isinstance(p, list):
        return [int(x) for x in p]
    elif isinstance(p, str):
        return [int(x) for x in p.strip('/').split('/') if x]
    return []


# ── Fallback maps (same as upstream scripts) ─────────────────────────────────
_ISOCORTEX_AREA_MAP = {
    'MOp': 'MO',  'MOs': 'MO',
    'SSp': 'SS',  'SSp-bfd': 'SS', 'SSp-tr': 'SS', 'SSp-ll': 'SS',
    'SSp-n': 'SS','SSp-m': 'SS',   'SSp-ul': 'SS', 'SSp-un': 'SS',
    'SSs': 'SS',
    'VISp': 'VIS','VISl': 'VIS',   'VISam': 'VIS', 'VISpm': 'VIS',
    'VISrl': 'VIS','VISa': 'VIS',  'VISal': 'VIS', 'VISli': 'VIS',
    'VISpor': 'VIS',
    'AUDp': 'AUD','AUDd': 'AUD',   'AUDv': 'AUD',  'AUDpo': 'AUD',
    'RSPagl': 'RSP','RSPd': 'RSP', 'RSPv': 'RSP',
    'ACAd': 'ACA','ACAv': 'ACA',
    'PL': 'PL',   'ILA': 'ILA',
    'ORBl': 'ORB','ORBm': 'ORB',   'ORBvl': 'ORB', 'ORBv': 'ORB',
    'AIp': 'AI',  'AIv': 'AI',     'AId': 'AI',
    'GU': 'GU',   'VISC': 'VISC',  'PTLp': 'PTLp',
    'TEa': 'TEa', 'ECT': 'ECT',    'PERI': 'PERI', 'FRP': 'FRP',
}

_OLF_FALLBACK = ['MOB','AOB','AON','TT','DP','PIR','NLOT','PAA','COA','OLF','EP','EPd','EPv']
_HPF_FALLBACK = ['CA1','CA2','CA3','DG','SUB','PRE','POST','PARA','FC','IG','HATA','APR','HPF','ProS']
_STR_FALLBACK = ['CP','ACB','FS','LSX','LS','CEA','BST','LA','BLA','BMA','AA','OT','MEA','SI','STR','isl','islm']

_LAYER_SUFFIXES = ['2/3', '6a', '6b', '1', '4', '5', '6']
_location_cache = {}
_hpf_nucleus_cache = {}
_thalamic_group_cache = {}


def _isocortex_area_fallback(base_area):
    if base_area in _ISOCORTEX_AREA_MAP:
        return _ISOCORTEX_AREA_MAP[base_area]
    for key, val in _ISOCORTEX_AREA_MAP.items():
        if base_area.startswith(key):
            return val
    return None


def _olf_hpf_str_fallback(base_area):
    loc = base_area.upper()
    for p in _OLF_FALLBACK:
        if loc == p.upper() or loc.startswith(p.upper()): return 'OLF'
    for p in _HPF_FALLBACK:
        if loc == p.upper() or loc.startswith(p.upper()): return 'HPF'
    for p in _STR_FALLBACK:
        if loc == p.upper() or loc.startswith(p.upper()): return 'STR'
    return None


def _try_strip_layer(location):
    for suffix in _LAYER_SUFFIXES:
        if str(location).endswith(suffix):
            base = location[:-len(suffix)]
            if len(base) >= 2:
                return base, suffix
    return location, None


def _query_allen(acronym):
    if not ALLEN_SDK_AVAILABLE:
        return None
    structs = structure_tree.get_structures_by_acronym([acronym])
    if not structs:
        return None
    path = _get_path(structs[0])
    if RT_ID in path:
        return 'Thalamus', None
    if STR_ID in path:
        return 'STR', None
    if ISOCORTEX_ID in path:
        iso_idx = path.index(ISOCORTEX_ID)
        if iso_idx + 1 < len(path):
            child = structure_tree.get_structures_by_id([path[iso_idx + 1]])
            if child:
                return 'Isocortex', child[0]['acronym']
        return 'Isocortex', acronym
    if OLF_ID in path: return 'OLF', None
    if HPF_ID in path: return 'HPF', None
    if TH_ID  in path: return 'Thalamus', None
    return None


def _classify_location(location):
    """Returns (base_area, layer_norm, major_region, isocortex_area)."""
    if location in _location_cache:
        return _location_cache[location]

    result = _query_allen(location)
    if result is not None:
        major_region, iso_area = result
        if major_region == 'Isocortex':
            base_full, layer_raw_full = _try_strip_layer(location)
            if base_full != location:
                r2 = _query_allen(base_full)
                iso2 = r2[1] if r2 and r2[0] == 'Isocortex' else iso_area
                out = (base_full, LAYER_NORMALIZE.get(layer_raw_full, layer_raw_full),
                       'Isocortex', iso2 or iso_area)
            else:
                out = (location, None, 'Isocortex', iso_area)
        else:
            out = (location, None, major_region, None)
        _location_cache[location] = out
        return out

    base, layer_raw = _try_strip_layer(location)
    if base != location:
        result = _query_allen(base)
        if result is not None:
            major_region, iso_area = result
            if major_region == 'Isocortex':
                out = (base, LAYER_NORMALIZE.get(layer_raw, layer_raw),
                       'Isocortex', iso_area)
            else:
                out = (base, None, major_region, None)
            _location_cache[location] = out
            return out

    mr = _olf_hpf_str_fallback(location)
    if mr:
        out = (location, None, mr, None)
        _location_cache[location] = out
        return out

    if base != location:
        mr = _olf_hpf_str_fallback(base)
        if mr:
            out = (base, None, mr, None)
            _location_cache[location] = out
            return out
        iso = _isocortex_area_fallback(base)
        if iso:
            out = (base, LAYER_NORMALIZE.get(layer_raw, layer_raw),
                   'Isocortex', iso)
            _location_cache[location] = out
            return out

    iso = _isocortex_area_fallback(location)
    if iso:
        out = (location, None, 'Isocortex', iso)
        _location_cache[location] = out
        return out

    out = (location, None, 'Other', None)
    _location_cache[location] = out
    return out


def get_hpf_nucleus(location):
    """For an HPF location, resolve the nucleus (CA1, CA2, CA3, DG, Subicular complex, etc.)."""
    if location in _hpf_nucleus_cache:
        return _hpf_nucleus_cache[location]

    result = location
    if ALLEN_SDK_AVAILABLE:
        structs = structure_tree.get_structures_by_acronym([location])
        if structs:
            path = _get_path(structs[0])
            if HPF_ID in path:
                hpf_idx = path.index(HPF_ID)
                if hpf_idx + 2 < len(path):
                    depth2_id = path[hpf_idx + 2]
                    if depth2_id == CA_ID and hpf_idx + 3 < len(path):
                        child = structure_tree.get_structures_by_id([path[hpf_idx + 3]])
                        if child:
                            result = child[0]['acronym']
                    else:
                        child = structure_tree.get_structures_by_id([depth2_id])
                        if child:
                            result = child[0]['acronym']
                elif hpf_idx + 1 < len(path):
                    child = structure_tree.get_structures_by_id([path[hpf_idx + 1]])
                    if child:
                        result = child[0]['acronym']

    result = HPF_NUCLEUS_MERGE.get(result, result)
    _hpf_nucleus_cache[location] = result
    return result


def get_thalamic_group(location):
    """Map a thalamic location to one of THALAMIC_GROUP_ORDER, or None."""
    if location in _thalamic_group_cache:
        return _thalamic_group_cache[location]

    if location in THALAMIC_GROUP_MAP:
        _thalamic_group_cache[location] = THALAMIC_GROUP_MAP[location]
        return THALAMIC_GROUP_MAP[location]

    result = None
    if ALLEN_SDK_AVAILABLE:
        structs = structure_tree.get_structures_by_acronym([location])
        if structs:
            path = _get_path(structs[0])
            if TH_ID in path:
                th_idx = path.index(TH_ID)
                for pid in path[th_idx + 1:]:
                    node = structure_tree.get_structures_by_id([pid])
                    if node:
                        acr = node[0]['acronym']
                        if acr in THALAMIC_GROUP_MAP:
                            result = THALAMIC_GROUP_MAP[acr]
                            break

    _thalamic_group_cache[location] = result
    return result


# ============================================================================
# DATA LOADING
# ============================================================================

def load_all_data(base_dir='.'):
    print("\nLoading data...")
    for candidate in [base_dir, '.', '..',
                      'output_prestim_burst', 'output_prestim_burst_new',
                      'output_prestim_burst_new4_silence50ms',
                      '../output_prestim_burst', '../output_prestim_burst_new',
                      '../output_prestim_burst_new4_silence50ms']:
        if (glob.glob(f'{candidate}/psilocybin/*_timecourse.csv') or
                glob.glob(f'{candidate}/saline/*_timecourse.csv')):
            base_dir = candidate
            break
    else:
        raise FileNotFoundError(f"No data found under: {base_dir}")
    print(f"  Using base_dir: {base_dir}")

    dfs = []
    for treatment, folder in [('Psilocybin', 'psilocybin'),
                               ('Saline',     'saline'),
                               ('Ketanserin', 'ketanserin')]:
        files = sorted(glob.glob(f'{base_dir}/{folder}/*_timecourse.csv'))
        print(f"  {treatment}: {len(files)} files")
        for f in files:
            df = pd.read_csv(f)
            df['session_id'] = os.path.basename(f).replace(
                '_prestim_burst_timecourse.csv', '')
            df['treatment'] = treatment
            dfs.append(df)

    data = pd.concat(dfs, ignore_index=True)
    data['period'] = data['bin'].apply(
        lambda x: 'Baseline' if 'baseline' in str(x) else 'Effect')

    data['time_bin_start'] = data['bin'].apply(
        lambda b: int(str(b).split('_')[1].split('-')[0]))

    for col in ['burst_fraction', 'burst_rate', 'mean_burst_size']:
        if col in data.columns:
            data[col] = data[col].where(data[col] > 0, np.nan)
    if 'total_spikes' in data.columns:
        data.loc[data['total_spikes'] == 0, 'burst_fraction'] = np.nan

    print(f"  Total rows: {len(data):,}")
    print(f"  Units     : {data[['session_id','unit_id']].drop_duplicates().shape[0]:,}")
    print(f"  Sessions  : {data['session_id'].nunique()}")
    return data


# ============================================================================
# ANNOTATION
# ============================================================================

def annotate(data):
    print("\nAnnotating locations...")
    data = data.copy()

    loc_cache = {}
    base_areas, layer_norms, major_regions, iso_areas = [], [], [], []
    for loc in data['location']:
        if loc not in loc_cache:
            loc_cache[loc] = _classify_location(loc)
        b, ln, mr, ia = loc_cache[loc]
        base_areas.append(b)
        layer_norms.append(ln)
        major_regions.append(mr)
        iso_areas.append(ia)

    data['base_area']      = base_areas
    data['layer_norm']     = layer_norms
    data['major_region']   = major_regions
    data['isocortex_area'] = iso_areas

    # Cortical row label  (e.g. "Prefrontal — L5")
    def _cortical_row(r):
        if r['major_region'] != 'Isocortex':
            return None
        grp = CORTICAL_AREA_TO_GROUP.get(r['isocortex_area'])
        if grp is None or r['layer_norm'] not in LAYER_ORDER:
            return None
        return _cortical_key(grp, r['layer_norm'])
    data['cortical_row'] = data.apply(_cortical_row, axis=1)

    # Thalamic group
    def _th_group(r):
        if r['major_region'] != 'Thalamus':
            return None
        for cand in [r['location'], r['base_area']]:
            g = get_thalamic_group(cand)
            if g is not None:
                return g
        return None
    data['thalamic_group'] = data.apply(_th_group, axis=1)

    # HPF nucleus
    def _hpf_nuc(r):
        if r['major_region'] != 'HPF':
            return None
        for cand in [r['location'], r['base_area']]:
            n = get_hpf_nucleus(cand)
            if n is not None:
                return n
        return None
    data['hpf_nucleus'] = data.apply(_hpf_nuc, axis=1)

    # Diagnostic counts
    print("\n  Unit counts by cortical_row:")
    cc = data.dropna(subset=['cortical_row']).drop_duplicates(['session_id', 'unit_id']).groupby(['cortical_row', 'treatment']).size().unstack(fill_value=0)
    if not cc.empty:
        print(cc.to_string())

    print("\n  Unit counts by thalamic_group:")
    tc = data.dropna(subset=['thalamic_group']).drop_duplicates(['session_id', 'unit_id']).groupby(['thalamic_group', 'treatment']).size().unstack(fill_value=0)
    if not tc.empty:
        print(tc.to_string())

    print("\n  Unit counts by hpf_nucleus:")
    hc = data.dropna(subset=['hpf_nucleus']).drop_duplicates(['session_id', 'unit_id']).groupby(['hpf_nucleus', 'treatment']).size().unstack(fill_value=0)
    if not hc.empty:
        print(hc.to_string())

    return data


# ============================================================================
# DELTA COMPUTATION  (per-unit pct change from baseline, 30–50 min window)
# ============================================================================

def compute_unit_deltas(data, metrics):
    print(f"\nComputing per-unit ABSOLUTE deltas ({WINDOW_LABEL})...")

    # Match heatmap exactly — group baseline by (session_id, unit_id)
    baseline_means = (data[data['period'] == 'Baseline']
                      .groupby(['session_id', 'unit_id'])[metrics]
                      .mean()
                      .rename(columns={m: f'baseline_{m}' for m in metrics})
                      .reset_index())
    enriched = data.merge(baseline_means, on=['session_id', 'unit_id'], how='left')

    delta_cols = []
    for m in metrics:
        bc = f'baseline_{m}'
        dc = f'{m}_delta'
        enriched[dc] = enriched[m] - enriched[bc]
        delta_cols.append(dc)

    effect = enriched[enriched['period'] == 'Effect'].copy()
    effect = effect[effect[delta_cols].notna().any(axis=1)].copy()

    meta_cols = ['session_id', 'unit_id', 'treatment', 'location',
                 'base_area', 'layer_norm', 'isocortex_area',
                 'major_region', 'cortical_row', 'thalamic_group', 'hpf_nucleus']
    meta_cols = [c for c in meta_cols if c in effect.columns]

    # Average across the effect window
    eff_win = (effect[effect['bin'].isin(EFFECT_WINDOW)]
               .groupby(['session_id', 'unit_id'])[delta_cols]
               .mean()
               .reset_index())

    meta = (effect[meta_cols]
            .drop_duplicates(subset=['session_id', 'unit_id', 'treatment']))

    merged = meta.merge(eff_win, on=['session_id', 'unit_id'], how='inner')
    merged = merged[merged[delta_cols].notna().any(axis=1)].copy()

    print(f"  Units with delta: {merged.groupby(['session_id', 'unit_id']).ngroups:,}")
    print(f"  Sessions        : {merged['session_id'].nunique()}")
    return merged


def compute_prop_bursting_deltas(data):
    """
    SESSION-LEVEL prop_bursting delta -- matches heatmap_hier_bootstrap.py /
    heatmap_hpf_thalamic_nuclei.py exactly: for each (region, session,
    treatment), the fraction of units in that session/region that are
    bursting in a given bin, NOT a per-unit fraction-of-time-bursting score.
    (The previous version of this function -- see RAW-PROP-1-style bug,
    logged as a distribution-script finding -- grouped by
    ['session_id', 'unit_id'] instead of by region+session, which computed a
    different, per-unit temporal-consistency statistic and fed it through the
    standard two-level hierarchical bootstrap instead of a session-only one.)

    unit_id is set equal to session_id afterward, exactly as the heatmap
    scripts do, so the existing hierarchical (session, unit) bootstrap
    machinery in compute_region_stats()/bootstrap_mean() collapses to pure
    session-only resampling without needing a separate bootstrap code path --
    resampling a group of size 1 (the session itself) with replacement is a
    no-op at the "unit" level.
    """
    print(f"\nComputing SESSION-LEVEL prop_bursting ABSOLUTE deltas ({WINDOW_LABEL})...")
    d = data.copy()
    # Match heatmap — exclude empty bins (rest_duration == 0) before computing
    # prop_bursting, so "no measurement" (NaN burst_fraction) is not turned
    # into a spurious "not bursting" (0) by fillna(0). See U-10.
    if 'rest_duration' in d.columns:
        d = d[d['rest_duration'] > 0].copy()
    d['prop_bursting'] = (d['burst_fraction'].fillna(0) > 0).astype(float)

    group_cols = [c for c in ['cortical_row', 'thalamic_group', 'hpf_nucleus'] if c in d.columns]

    chunks = []
    for gc in group_cols:
        sub = d[d[gc].notna()].copy()
        if len(sub) == 0:
            continue

        baseline = (sub[sub['period'] == 'Baseline']
                    .groupby([gc, 'session_id', 'treatment'])['prop_bursting']
                    .mean()
                    .reset_index()
                    .rename(columns={'prop_bursting': 'baseline_prop'}))

        eff = sub[(sub['period'] == 'Effect') & (sub['bin'].isin(EFFECT_WINDOW))].copy()
        eff_sess = (eff.groupby([gc, 'session_id', 'treatment'])['prop_bursting']
                    .mean()
                    .reset_index()
                    .rename(columns={'prop_bursting': 'effect_prop'}))

        merged_gc = eff_sess.merge(baseline, on=[gc, 'session_id', 'treatment'], how='left')
        merged_gc['prop_bursting_delta'] = merged_gc['effect_prop'] - merged_gc['baseline_prop']
        merged_gc = merged_gc.dropna(subset=['prop_bursting_delta'])
        merged_gc['unit_id'] = merged_gc['session_id']  # session-only bootstrap trick, matches heatmap

        for other in group_cols:
            if other != gc:
                merged_gc[other] = np.nan

        chunks.append(merged_gc[['session_id', 'unit_id', 'treatment'] + group_cols + ['prop_bursting_delta']])

    merged = pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(
        columns=['session_id', 'unit_id', 'treatment'] + group_cols + ['prop_bursting_delta'])

    print(f"  Session-region rows: {len(merged):,}, Sessions: {merged['session_id'].nunique()}")
    return merged


# ============================================================================
# HIERARCHICAL BOOTSTRAP  (session → unit)
# ============================================================================

def _sample_hier(df, metric, levels, num_samples, nboots):
    if len(levels) == 0:
        sums   = np.zeros(nboots)
        counts = np.zeros(nboots)
        for i in range(nboots):
            if num_samples[i] > 0:
                n = len(df) * int(num_samples[i])
                sums[i]   = df[metric].sample(n=n, replace=True).sum()
                counts[i] = n
        return sums, counts
    items   = df[levels[0]].unique()
    samples = np.zeros((len(items), nboots))
    for i in range(nboots):
        chosen = Counter(np.random.choice(
            items, size=len(items) * int(num_samples[i])))
        for idx, item in enumerate(items):
            samples[idx, i] = chosen[item]
    sums   = np.zeros(nboots)
    counts = np.zeros(nboots)
    for idx, item in enumerate(items):
        temp = df[df[levels[0]] == item]
        ts, tc = _sample_hier(temp, metric, levels[1:], samples[idx, :], nboots)
        sums   += ts
        counts += tc
    return sums, counts


def bootstrap_mean(df, metric, levels, nboots=1000):
    sums, counts = _sample_hier(df, metric, levels, [1] * nboots, nboots)
    valid  = counts > 0
    result = np.full(nboots, np.nan)
    result[valid] = sums[valid] / counts[valid]
    return result


# ============================================================================
# REGION-LEVEL STATS  (returns stats DataFrame + raw bootstrap arrays)
# ============================================================================

def compute_region_stats(deltas, group_col, regions, conditions,
                         delta_col, nboots=1000):
    levels = ['session_id', 'unit_id']  # hierarchical
    drug_cond, ctrl_cond = conditions
    records = []
    boots_by_region = {}

    for region in tqdm(regions, desc=f"{delta_col}"):
        rdata = deltas[deltas[group_col] == region].copy()
        if len(rdata) == 0:
            continue

        cond_boots = {}
        for cond in conditions:
            cd = rdata[rdata['treatment'] == cond].dropna(subset=[delta_col])
            if len(cd) < 2:
                continue
            boot = bootstrap_mean(
                cd[['session_id', 'unit_id', delta_col]],
                metric=delta_col, levels=levels, nboots=nboots)
            cond_boots[cond] = boot

        if drug_cond not in cond_boots or ctrl_cond not in cond_boots:
            continue

        db    = cond_boots[drug_cond]
        cb    = cond_boots[ctrl_cond]
        valid = ~(np.isnan(db) | np.isnan(cb))
        if valid.sum() < 10:
            continue

        diff   = db[valid] - cb[valid]
        effect = np.mean(diff)
        p1     = np.mean(diff < 0) if effect >= 0 else np.mean(diff > 0)
        p_val  = min(2 * p1, 1.0)

        dv = rdata[rdata['treatment'] == drug_cond][delta_col].dropna()
        cv = rdata[rdata['treatment'] == ctrl_cond][delta_col].dropna()
        if len(dv) > 1 and len(cv) > 1:
            pooled_sd = np.sqrt((dv.std(ddof=1)**2 + cv.std(ddof=1)**2) / 2)
            cohens_d  = (dv.mean() - cv.mean()) / pooled_sd if pooled_sd > 0 else np.nan
        else:
            cohens_d = np.nan

        row = {group_col: region, 'p_value': p_val, 'cohens_d': cohens_d}
        for cond in conditions:
            if cond in cond_boots:
                b = cond_boots[cond]
                row[f'{cond}_mean']  = np.nanmean(b)
                row[f'{cond}_ci_lo'] = np.nanpercentile(b, 2.5)
                row[f'{cond}_ci_hi'] = np.nanpercentile(b, 97.5)
                cd_cond = rdata[rdata['treatment'] == cond].dropna(subset=[delta_col])
                row[f'{cond}_n_units']    = int(cd_cond.groupby(['session_id', 'unit_id']).ngroups)
                row[f'{cond}_n_sessions'] = int(cd_cond['session_id'].nunique())
        records.append(row)
        boots_by_region[region] = cond_boots

    return pd.DataFrame(records), boots_by_region


# ============================================================================
# FDR (Benjamini–Hochberg)
# ============================================================================

def apply_fdr(stats_df):
    if len(stats_df) == 0 or 'p_value' not in stats_df.columns:
        return stats_df
    p = stats_df['p_value'].values.copy()
    n = len(p)
    try:
        from scipy.stats import false_discovery_control
        p_adj = false_discovery_control(np.clip(p, 1e-10, 1.0), method='bh')
    except Exception:
        si    = np.argsort(p)
        sp    = p[si]
        p_adj = np.minimum(1.0, sp * n / np.arange(1, n + 1))
        for i in range(n - 2, -1, -1):
            p_adj[i] = min(p_adj[i], p_adj[i + 1])
        reord     = np.empty(n)
        reord[si] = p_adj
        p_adj     = reord
    out = stats_df.copy()
    out['p_fdr'] = p_adj
    return out


def stars(p):
    if   p < 0.001: return '***'
    elif p < 0.01:  return '**'
    elif p < 0.05:  return '*'
    return ''


# ============================================================================
# FILTERING
# ============================================================================

def filter_regions(deltas, group_col, conditions, min_units=10, min_sessions=3, verbose=True):
    valid = []
    rejected = []
    for region in deltas[group_col].dropna().unique():
        rdata  = deltas[deltas[group_col] == region]
        passes = True
        reasons = []
        for cond in conditions:
            cd = rdata[rdata['treatment'] == cond]
            n_u = cd.groupby(['session_id', 'unit_id']).ngroups
            n_s = cd['session_id'].nunique()
            if n_u < min_units or n_s < min_sessions:
                passes = False
                reasons.append(f'{cond}: {n_u}u/{n_s}s')
        if passes:
            valid.append(region)
        else:
            rejected.append((region, '; '.join(reasons)))
    if verbose:
        print(f"  filter [{'+'.join(conditions)}] min_units={min_units}, min_sessions={min_sessions}:")
        print(f"    KEEP ({len(valid)}): {valid}")
        if rejected:
            print(f"    DROP ({len(rejected)}):")
            for region, reason in rejected:
                print(f"      {region}  →  {reason}")
    return valid


# ============================================================================
# PLOT — half-violin + point + CI  (matches screenshot style)
# ============================================================================

def _plot_half_violin(ax, x, density_max, y_base, direction, color,
                      half_height=0.38, alpha=0.35, fill=False):
    """Draw half-violin from y_base in `direction` (+1 up / -1 down).
    fill=True  → filled polygon
    fill=False → outline curve only (closed back to the baseline)
    """
    y_top = y_base + direction * density_max * half_height
    if fill:
        xs = np.concatenate([x, x[::-1]])
        ys = np.concatenate([np.full_like(x, y_base), y_top[::-1]])
        ax.fill(xs, ys, color=color, alpha=alpha, lw=0, zorder=2)
        ax.plot(x, y_top, color=color, lw=0.9, alpha=0.85, zorder=2)
    else:
        # Outline only — curve from baseline up through density and back
        xs = np.concatenate([[x[0]], x, [x[-1]]])
        ys = np.concatenate([[y_base], y_top, [y_base]])
        ax.plot(xs, ys, color=color, lw=1.1, alpha=0.95, zorder=2)


def forest_panel_with_violins(ax, stats_df, boots_dict, group_col, regions,
                              drug_cond, ctrl_cond, label_map,
                              sig_col='p_fdr'):
    drug_color = CONDITION_COLORS[drug_cond]
    ctrl_color = CONDITION_COLORS[ctrl_cond]
    point_off  = 0.14
    half_h     = 0.38
    n          = len(regions)
    x_all      = []
    sig_annots = []

    for i, region in enumerate(regions):
        rows = stats_df[stats_df[group_col] == region]
        if len(rows) == 0:
            continue
        row   = rows.iloc[0]
        boots = boots_dict.get(region, {})

        for cond, direction, color, yoff_sign in [
            # direction is in DATA y-space; ax.invert_yaxis() flips the
            # display, so visually drug ends up BELOW the row label and
            # control above — matching the screenshot reference.
            (drug_cond, +1, drug_color, +1),
            (ctrl_cond, -1, ctrl_color, -1),
        ]:
            # Half-violin from bootstrap distribution
            b = boots.get(cond)
            if b is not None:
                bv = b[~np.isnan(b)]
                if len(bv) > 5 and np.std(bv) > 1e-10:
                    try:
                        kde = gaussian_kde(bv)
                        lo, hi = np.percentile(bv, [0.5, 99.5])
                        pad = (hi - lo) * 0.05 + 1e-9
                        x_grid = np.linspace(lo - pad, hi + pad, 200)
                        dens   = kde(x_grid)
                        if dens.max() > 0:
                            dens = dens / dens.max()
                            _plot_half_violin(
                                ax, x_grid, dens, y_base=i,
                                direction=direction, color=color,
                                half_height=half_h)
                            x_all.extend([x_grid[0], x_grid[-1]])
                    except Exception:
                        pass

            # Point only (no horizontal CI bar — the violin shows the spread)
            m_key  = f'{cond}_mean'
            if m_key in row.index and not pd.isna(row[m_key]):
                m  = row[m_key]
                y_pt = i + yoff_sign * point_off
                ax.scatter(m, y_pt, color=color, s=42, zorder=5,
                           edgecolors='white', linewidths=0.5)
                # Still gather x extent from the CI for axis scaling
                lo_key = f'{cond}_ci_lo'
                hi_key = f'{cond}_ci_hi'
                if lo_key in row.index and not pd.isna(row.get(lo_key)):
                    x_all.extend([row[lo_key], row[hi_key]])

        p_unc  = row.get('p_value', 1.0)
        p_corr = row.get(sig_col, 1.0)
        d_val  = row.get('cohens_d', np.nan)
        star_corr = stars(p_corr if not pd.isna(p_corr) else 1.0)
        star_unc  = stars(p_unc  if not pd.isna(p_unc)  else 1.0)
        sig_annots.append((i, star_corr, star_unc, p_corr, d_val))

    ax.axvline(0, color='black', lw=1.0, ls='--', zorder=1, alpha=0.7)
    for i in range(n):
        ax.axhline(i, color='grey', lw=0.4, alpha=0.3, zorder=0)

    ax.set_yticks(np.arange(n))
    ax.set_yticklabels([label_map.get(r, r) for r in regions], fontsize=9)
    ax.set_ylim(-0.6, n - 0.4)
    ax.invert_yaxis()
    ax.spines[['top', 'right']].set_visible(False)
    return x_all, sig_annots


def apply_xlim_and_annotations(ax, xlim, sig_annots, drug_color):
    xmin, xmax = xlim
    span = max(abs(xmax - xmin), 1e-6)
    pad  = max(span * 0.05, abs(xmax) * 0.03, 0.001)
    ax.set_xlim(xmin - pad, xmax + pad * 14)  # extra room for annotations

    xr   = ax.get_xlim()[1]
    xl   = ax.get_xlim()[0]
    xpos = xr - (xr - xl) * 0.005

    for (i, star_corr, star_unc, p_corr, d_val) in sig_annots:
        # Tier 1: FDR-corrected star (bold, large) if present
        # Tier 2: uncorrected star only (lighter, smaller, in parens) if uncorr<.05 but FDR>=.05
        if star_corr:
            marker = star_corr
            fw, fs, alpha = 'bold', 10, 1.0
        elif star_unc:
            marker = f'({star_unc})'    # parens = uncorrected only
            fw, fs, alpha = 'normal', 7.5, 0.85
        else:
            marker = ''
            fw, fs, alpha = 'normal', 7, 0.7
        bits = []
        if marker:
            bits.append(marker)
        if not pd.isna(p_corr):
            bits.append(f'p_fdr={p_corr:.3f}')
        if not pd.isna(d_val):
            bits.append(f'd={d_val:.2f}')
        if bits:
            ax.text(xpos, i, '\n'.join(bits), ha='right', va='center',
                    fontsize=fs, color=drug_color, fontweight=fw, alpha=alpha)


# ============================================================================
# FIGURE BUILDER
# ============================================================================

def make_figure(deltas, group_col, regions, comparisons, delta_col,
                title_main, xlabel_str, out_path, label_map,
                nboots=1000, min_units=1, min_sessions=1,
                stats_csv_path=None):
    n_comp = len(comparisons)
    n_reg  = len(regions)
    panel_h = max(4.0, n_reg * 0.5 + 1.5)
    fig, axes = plt.subplots(
        n_comp, 1,
        figsize=(8.5, panel_h * n_comp + 1.2),
        constrained_layout=True)
    if n_comp == 1:
        axes = np.array([axes])

    all_stats     = {}
    all_boots     = {}
    all_regions   = {}
    x_all_global  = []
    stats_records = []

    # --- first pass: stats + bootstraps + xlim gather ---
    for row_idx, (drug_cond, ctrl_cond) in enumerate(comparisons):
        conds = [drug_cond, ctrl_cond]
        # prop_bursting has unit_id == session_id (see compute_prop_bursting_deltas),
        # so its "units" and "sessions" counts are numerically identical -- applying
        # the normal min_units threshold on top of min_sessions would silently
        # require e.g. >=10 sessions, which Ketanserin (8 sessions total in the
        # whole dataset) could never pass for any region. Matches the heatmap
        # scripts, which gate region validity once from the unit-level data and
        # never re-gate prop_bursting by its own session count.
        eff_min_units = 1 if delta_col == 'prop_bursting_delta' else min_units
        valid = set(filter_regions(deltas, group_col, conds, eff_min_units, min_sessions))
        plot_regions = [r for r in regions if r in valid]
        all_regions[row_idx] = plot_regions

        if not plot_regions:
            continue

        stats, boots = compute_region_stats(
            deltas, group_col, plot_regions, conds,
            delta_col, nboots=nboots)
        stats = apply_fdr(stats)  # BH across rows in this panel
        all_stats[row_idx] = stats
        all_boots[row_idx] = boots

        # CSV log
        slog = stats.copy()
        slog['comparison'] = f'{drug_cond}_vs_{ctrl_cond}'
        slog['delta_col']  = delta_col
        stats_records.append(slog)

        for cond in conds:
            for key in [f'{cond}_ci_lo', f'{cond}_ci_hi']:
                if key in stats.columns:
                    x_all_global.extend(stats[key].dropna().tolist())
        # Also include violin support so x-axis fits the distributions
        for region, cb in boots.items():
            for cond, b in cb.items():
                bv = b[~np.isnan(b)]
                if len(bv) > 0:
                    x_all_global.extend(np.percentile(bv, [1, 99]).tolist())

    xs = [x for x in x_all_global if np.isfinite(x)]
    shared_xlim = (min(xs), max(xs)) if xs else (-1, 1)

    # --- second pass: draw ---
    for row_idx, (drug_cond, ctrl_cond) in enumerate(comparisons):
        ax = axes[row_idx]
        plot_regions = all_regions.get(row_idx, [])

        if not plot_regions:
            ax.text(0.5, 0.5, 'No data\n(filter threshold not met)',
                    transform=ax.transAxes, ha='center', va='center',
                    fontsize=11, color='grey')
            ax.set_axis_off()
            continue

        stats = all_stats.get(row_idx, pd.DataFrame())
        boots = all_boots.get(row_idx, {})

        x_vals, sig_annots = forest_panel_with_violins(
            ax, stats, boots, group_col, plot_regions,
            drug_cond, ctrl_cond, label_map=label_map, sig_col='p_fdr')

        apply_xlim_and_annotations(
            ax, shared_xlim, sig_annots, CONDITION_COLORS[drug_cond])

        ax.set_title(
            f'{drug_cond} vs {ctrl_cond}  '
            f'[hierarchical bootstrap (session → unit), FDR corrected]',
            fontsize=11, fontweight='bold', pad=6)
        ax.set_xlabel(xlabel_str, fontsize=10)

    metric_key  = '_'.join(delta_col.split('_')[2:])
    metric_nice = METRIC_LABELS.get(metric_key, metric_key)
    fig.suptitle(
        f'{title_main}  |  {metric_nice}  |  {WINDOW_LABEL}',
        fontsize=12.5, fontweight='bold', y=1.005)

    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {out_path}")

    if stats_csv_path and stats_records:
        df = pd.concat(stats_records, ignore_index=True)
        os.makedirs(os.path.dirname(stats_csv_path) or '.', exist_ok=True)
        df.to_csv(stats_csv_path, index=False)
        print(f"  Saved: {stats_csv_path}")


# ============================================================================
# RUN ONE PLOT GROUP (cortical / thalamic / hpf) — all 4 metrics
# ============================================================================

def run_plot_group(deltas, prop_deltas, *, group_col, row_order, label_map,
                   title_main, out_subdir, nboots,
                   min_units=1, min_sessions=1):
    metrics = ['burst_fraction', 'burst_rate', 'mean_burst_size']

    # Only keep rows whose row label appears in deltas (in canonical order)
    present_norm = set(deltas[group_col].dropna().unique())
    rows_norm    = [r for r in row_order if r in present_norm]
    print(f"\n[{title_main}] rows present (delta metrics): {rows_norm}")

    for m in metrics:
        delta_col = f'{m}_delta'
        if delta_col not in deltas.columns:
            print(f"  Skipping {delta_col} — column missing")
            continue
        out_path  = os.path.join(out_subdir, f'{m}_delta.png')
        stats_csv = os.path.join(out_subdir, f'{m}_delta_stats.csv')
        xlabel    = f'Δ {METRIC_LABELS[m]} (absolute change from baseline)'
        print(f"\n=== {title_main} | {m} ===")
        make_figure(
            deltas, group_col, rows_norm, COMPARISONS, delta_col,
            title_main=title_main, xlabel_str=xlabel,
            out_path=out_path, label_map=label_map,
            nboots=nboots, min_units=min_units, min_sessions=min_sessions,
            stats_csv_path=stats_csv)

    # prop_bursting
    present_prop = set(prop_deltas[group_col].dropna().unique())
    rows_prop    = [r for r in row_order if r in present_prop]
    print(f"\n[{title_main}] rows present (prop_bursting): {rows_prop}")
    delta_col = 'prop_bursting_delta'
    out_path  = os.path.join(out_subdir, 'prop_bursting_delta.png')
    stats_csv = os.path.join(out_subdir, 'prop_bursting_delta_stats.csv')
    xlabel    = 'Δ Prop. Bursting Units (absolute change from baseline)'
    print(f"\n=== {title_main} | prop_bursting ===")
    make_figure(
        prop_deltas, group_col, rows_prop, COMPARISONS, delta_col,
        title_main=title_main, xlabel_str=xlabel,
        out_path=out_path, label_map=label_map,
        nboots=nboots, min_units=min_units, min_sessions=min_sessions,
        stats_csv_path=stats_csv)


# ============================================================================
# EXCEL EXPORT  (both uncorrected p_value and FDR-corrected p_fdr, side by side,
# with the same *** / (***) marker convention used on the figures)
# ============================================================================

_XL_FIXED_COLS = {'p_value', 'cohens_d', 'p_fdr', 'comparison', 'delta_col'}
_XL_COND_SUFFIXES = ('_mean', '_ci_lo', '_ci_hi', '_n_units', '_n_sessions')

_XL_COLUMN_ORDER = [
    'panel', 'window_min', 'metric', 'region', 'comparison', 'drug', 'control',
    'drug_mean', 'drug_ci_lo', 'drug_ci_hi',
    'control_mean', 'control_ci_lo', 'control_ci_hi',
    'cohens_d',
    'p_FDR', 'significance', 'sig_FDR', 'survives_FDR',
    'p_uncorrected', 'sig_uncorrected',
    'drug_n_units', 'drug_n_sessions', 'control_n_units', 'control_n_sessions',
]
_XL_INT = {'drug_n_units', 'drug_n_sessions', 'control_n_units', 'control_n_sessions'}
_XL_P = {'p_uncorrected', 'p_FDR'}
_XL_3DP = {'drug_mean', 'drug_ci_lo', 'drug_ci_hi',
           'control_mean', 'control_ci_lo', 'control_ci_hi'}
_XL_2DP = {'cohens_d'}


def _xl_marker(p_fdr, p_unc):
    """Plot-style marker: plain stars if FDR-significant, parenthesized stars if
    significant uncorrected but NOT surviving FDR (matches the figures)."""
    s_fdr = stars(p_fdr if not pd.isna(p_fdr) else 1.0)
    s_unc = stars(p_unc if not pd.isna(p_unc) else 1.0)
    return (s_fdr, s_unc, s_fdr if s_fdr else (f'({s_unc})' if s_unc else ''))


def _xl_region_col(df):
    for c in df.columns:
        if c in _XL_FIXED_COLS:
            continue
        if any(c.endswith(s) for s in _XL_COND_SUFFIXES):
            continue
        return c
    return None


def _xl_load_csv(csv_path, win_tag):
    df = pd.read_csv(csv_path)
    if df.empty:
        return None
    region_col = _xl_region_col(df)
    panel = os.path.basename(os.path.dirname(csv_path))
    rows = []
    for _, r in df.iterrows():
        comparison = str(r.get('comparison', ''))
        drug, ctrl = (comparison.split('_vs_', 1) + [''])[:2] if '_vs_' in comparison else (comparison, '')
        delta_col = str(r.get('delta_col', ''))
        metric = delta_col[:-6] if delta_col.endswith('_delta') else delta_col
        p_unc = r.get('p_value', np.nan)
        p_fdr = r.get('p_fdr', np.nan)
        s_fdr, s_unc, marker = _xl_marker(p_fdr, p_unc)
        out = {
            'panel': panel,
            'window_min': win_tag.replace('_', '-') if win_tag else '',
            'metric': metric,
            'region': r.get(region_col, '') if region_col else '',
            'comparison': comparison.replace('_', ' '),
            'drug': drug, 'control': ctrl,
            'cohens_d': r.get('cohens_d', np.nan),
            'p_FDR': p_fdr, 'significance': marker, 'sig_FDR': s_fdr,
            'survives_FDR': (not pd.isna(p_fdr)) and p_fdr < 0.05,
            'p_uncorrected': p_unc, 'sig_uncorrected': s_unc,
        }
        for role, cond in (('drug', drug), ('control', ctrl)):
            for suf, key in (('_mean', 'mean'), ('_ci_lo', 'ci_lo'), ('_ci_hi', 'ci_hi'),
                             ('_n_units', 'n_units'), ('_n_sessions', 'n_sessions')):
                out[f'{role}_{key}'] = r.get(f'{cond}{suf}', np.nan)
        rows.append(out)
    return pd.DataFrame(rows)


def _xl_style(ws, df):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='305496')
    body = Font(name='Arial')
    bold = Font(name='Arial', bold=True)
    thin = Border(bottom=Side(style='thin', color='D9D9D9'))
    idx = {c: i + 1 for i, c in enumerate(df.columns)}
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(1, j, col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(df.columns))}1'
    for i in range(len(df)):
        rr = i + 2
        surv = bool(df.iloc[i].get('survives_FDR', False))
        for col in df.columns:
            j = idx[col]; val = df.iloc[i][col]
            cell = ws.cell(rr, j)
            if pd.isna(val):
                cell.value = None
            elif col == 'survives_FDR':
                cell.value = 'YES' if bool(val) else ''
            else:
                cell.value = val
            cell.font = bold if (surv and col in ('region', 'p_FDR', 'significance', 'sig_FDR')) else body
            cell.border = thin
            if col in _XL_INT:
                cell.number_format = '0'; cell.alignment = Alignment(horizontal='center')
            elif col in _XL_P or col in _XL_3DP:
                cell.number_format = '0.000'; cell.alignment = Alignment(horizontal='center')
            elif col in _XL_2DP:
                cell.number_format = '0.00'; cell.alignment = Alignment(horizontal='center')
            elif col in ('significance', 'sig_FDR', 'sig_uncorrected', 'survives_FDR', 'window_min'):
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
    for col in df.columns:
        w = max(len(str(col)), *(len(str(v)) for v in df[col].astype(str))) + 2
        ws.column_dimensions[get_column_letter(idx[col])].width = min(max(w, 8), 40)


def export_stats_to_excel(out_dir, xlsx_path, win_tag=''):
    """Collect every *_delta_stats.csv under out_dir into one paper-ready xlsx."""
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  openpyxl not installed - skipping Excel export "
              "(pip install openpyxl)")
        return
    csvs = sorted(glob.glob(os.path.join(out_dir, '**', '*_delta_stats.csv'), recursive=True))
    if not csvs:
        print(f"  No *_delta_stats.csv under {out_dir} - Excel not written")
        return
    frames = []
    for p in csvs:
        try:
            f = _xl_load_csv(p, win_tag)
            if f is not None and not f.empty:
                f.insert(0, 'source_file', os.path.relpath(p, out_dir))
                frames.append(f)
        except Exception as e:
            print(f"  WARNING: skipped {p} ({e})")
    if not frames:
        return
    combined = pd.concat(frames, ignore_index=True)
    ordered = ['source_file'] + [c for c in _XL_COLUMN_ORDER if c in combined.columns]
    ordered += [c for c in combined.columns if c not in ordered]
    combined = combined[ordered]
    keys = [k for k in ['panel', 'metric', 'comparison', 'region'] if k in combined.columns]
    combined = combined.sort_values(keys, kind='stable').reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'All_stats'
    _xl_style(ws, combined)
    for metric, sub in combined.groupby('metric', sort=True):
        title = re.sub(r'[^A-Za-z0-9_]', '_', str(metric))[:28] or 'metric'
        ws2 = wb.create_sheet(title)
        _xl_style(ws2, sub.reset_index(drop=True))
    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)) or '.', exist_ok=True)
    wb.save(xlsx_path)
    n = int(combined['survives_FDR'].sum()) if 'survives_FDR' in combined else 0
    print(f"  Saved: {xlsx_path}  ({len(combined)} rows, {n} surviving FDR)")


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Distribution plots (half-violin + point/CI) — configurable window',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    parser.add_argument('--data_dir', default='output_prestim_burst_new',
                        help='Root data dir (psilocybin/, saline/, ketanserin/ subfolders)')
    parser.add_argument('--out_dir',  default='distribution_plots_40_60')
    parser.add_argument('--excel_out', default=None,
                        help='Excel output path (default: <out_dir>/cortex_thalamus_hpf_stats_<window>.xlsx)')
    parser.add_argument('--nboots',   type=int, default=10000)
    parser.add_argument('--seed',     type=int, default=42)
    parser.add_argument('--min_units',    type=int, default=10,
                        help='Min units PER CONDITION per region (default: 10, matches heatmap)')
    parser.add_argument('--min_sessions', type=int, default=3,
                        help='Min sessions PER CONDITION per region (default: 3, matches heatmap)')
    parser.add_argument('--window_start', type=int, default=40,
                        help='Effect window start in minutes post-injection (default: 40)')
    parser.add_argument('--window_end',   type=int, default=60,
                        help='Effect window end in minutes post-injection (default: 60)')
    args = parser.parse_args()

    # Build effect window from CLI args (in 5-min bins)
    global EFFECT_WINDOW, WINDOW_LABEL
    bin_edges = list(range(args.window_start, args.window_end, 5))
    EFFECT_WINDOW = {f'effect_{s}-{s+5}' for s in bin_edges}
    WINDOW_LABEL  = f'{args.window_start}-{args.window_end} min post-injection'
    print(f"Effect window: {WINDOW_LABEL}  →  bins: {sorted(EFFECT_WINDOW)}")

    np.random.seed(args.seed)

    metrics = ['burst_fraction', 'burst_rate', 'mean_burst_size']

    raw     = load_all_data(args.data_dir)
    data    = annotate(raw)
    deltas  = compute_unit_deltas(data, metrics)
    prop_d  = compute_prop_bursting_deltas(data)

    win_tag = f'{args.window_start}_{args.window_end}'

    # ── Plot 1: Cortical layer rows ──────────────────────────────────────────
    # Re-seed before each panel so every panel's bootstrap is independent of the
    # ones before it and reproducible regardless of processing order.
    np.random.seed(args.seed)
    run_plot_group(
        deltas, prop_d,
        group_col='cortical_row',
        row_order=CORTICAL_ROW_ORDER,
        label_map=CORTICAL_ROW_LABELS,
        title_main='Cortical Layers (Prefrontal / Somatosensory / Visual / RSP)',
        out_subdir=os.path.join(args.out_dir, f'cortical_layers_{win_tag}'),
        nboots=args.nboots,
        min_units=args.min_units, min_sessions=args.min_sessions)

    # ── Plot 2: Thalamic functional groups ───────────────────────────────────
    np.random.seed(args.seed)
    run_plot_group(
        deltas, prop_d,
        group_col='thalamic_group',
        row_order=THALAMIC_GROUP_ORDER,
        label_map={k: k for k in THALAMIC_GROUP_ORDER},
        title_main='Thalamic Functional Groups',
        out_subdir=os.path.join(args.out_dir, f'thalamic_groups_{win_tag}'),
        nboots=args.nboots,
        min_units=args.min_units, min_sessions=args.min_sessions)

    # ── Plot 3: Hippocampal nuclei ───────────────────────────────────────────
    # Re-seed here so the hpf bootstrap is independent of the cortical/thalamic
    # panels above and reproduces the standalone hippocampus script exactly
    # (which bootstraps hpf first, right after seeding).
    np.random.seed(args.seed)
    run_plot_group(
        deltas, prop_d,
        group_col='hpf_nucleus',
        row_order=HPF_ROW_ORDER,
        label_map={k: k for k in HPF_ROW_ORDER},
        title_main='Hippocampal Nuclei',
        out_subdir=os.path.join(args.out_dir, f'hpf_nuclei_{win_tag}'),
        nboots=args.nboots,
        min_units=args.min_units, min_sessions=args.min_sessions)

    # ── Excel: both uncorrected + FDR p-values, with the *** / (***) markers ──
    xlsx_path = args.excel_out or os.path.join(args.out_dir, f'cortex_thalamus_hpf_stats_{win_tag}.xlsx')
    print("\nExporting stats to Excel...")
    export_stats_to_excel(args.out_dir, xlsx_path, win_tag=win_tag)

    print(f"\n{'='*70}")
    print(f"Done. Outputs in: {args.out_dir}/")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
